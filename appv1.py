import streamlit as st
import pandas as pd
from google import genai
import io
import json
import openpyxl

# 페이지 설정
st.set_page_config(page_title="Excel Merger AI (Pro)", layout="wide")

st.title("📊 재무제표 통합 및 전체 계정 매핑 (Pro)")
st.markdown("여러 엑셀 파일을 업로드하면 **계정별로 상세 매핑**하여 **완결된 재무제표**를 생성합니다.")
st.markdown("ℹ️ **축약 없이 모든 계정을 나열**하며, 숨겨진 데이터는 제외합니다.")

# --- API Key Session State 관리 ---
if 'api_key' not in st.session_state:
    st.session_state.api_key = ''

with st.sidebar:
    st.header("설정")
    api_input = st.text_input(
        "Gemini API Key", 
        type="password", 
        placeholder="여기에 키를 입력하세요",
        value=st.session_state.api_key
    )
    if api_input:
        st.session_state.api_key = api_input

    if not st.session_state.api_key:
        st.warning("먼저 API 키를 입력해주세요.")

# --- 정밀 파싱 함수 (숨김 항목 제외) ---
def load_excel_visible_only(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    all_dfs = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state == 'hidden' or ws.sheet_state == 'veryHidden':
            continue
        
        visible_data = []
        for row_idx, row_cells in enumerate(ws.iter_rows(values_only=True), 1):
            if ws.row_dimensions[row_idx].hidden:
                continue
            if not any(row_cells):
                continue
            visible_data.append(row_cells)
        
        if visible_data:
            headers = visible_data[0]
            clean_headers = [str(h) if h is not None else f"Unnamed_{i}" for i, h in enumerate(headers)]
            df = pd.DataFrame(visible_data[1:], columns=clean_headers)
            df['Source'] = f"{file.name} - {sheet_name}"
            all_dfs.append(df)
            
    return all_dfs

# --- 메인 로직 ---
uploaded_files = st.file_uploader("엑셀 파일을 드래그하거나 선택하세요", accept_multiple_files=True, type=['xlsx', 'xls'])

if uploaded_files and st.session_state.api_key:
    if st.button("상세 재무제표 생성 시작"):
        all_data = []
        progress_text = st.empty()
        
        try:
            # 1. 파일 읽기
            progress_text.text("📂 엑셀 파일 정밀 파싱 중...")
            for file in uploaded_files:
                dfs = load_excel_visible_only(file)
                all_data.extend(dfs)
            
            if not all_data:
                st.error("처리할 데이터가 없습니다.")
            else:
                merged_df = pd.concat(all_data, ignore_index=True)
                st.success(f"✅ 데이터 로드 완료: 총 {len(uploaded_files)}개 파일, {len(merged_df)}행")
                
                with st.expander("병합된 원본 데이터 확인"):
                    st.dataframe(merged_df)

                # 2. Gemini AI 분석 (상세 매핑 요청)
                progress_text.text("🤖 AI가 계정별 매핑 작업을 수행 중입니다... (데이터 양에 따라 시간이 소요됩니다)")
                
                # 상세 분석을 위해 데이터 길이 제한을 좀 더 늘림 (토큰 허용 범위 내)
                csv_data = merged_df.to_csv(index=False)
                if len(csv_data) > 100000: # 10만 자로 제한 (Gemini Pro/Flash 계열은 컨텍스트가 큼)
                    csv_data = csv_data[:100000] + "\n...(이후 데이터 생략됨)"

                client = genai.Client(api_key=st.session_state.api_key)
                
                # --- [핵심 수정] 프롬프트: 축약 금지 및 상세 매핑 요청 ---
                prompt = f"""
                당신은 기업의 수석 회계사(Chief Accountant)입니다. 
                아래 제공된 재무 데이터를 바탕으로 "상세 연도별 비교 재무제표(Detailed Comparative Financial Statement)"를 작성하십시오.

                [강력한 제약사항 - 절대 준수]
                1. **절대 계정을 축약하거나 임의로 합치지 마십시오.** (Do not summarize).
                2. 원본 데이터에 있는 **모든 세부 계정 과목(Account Item)**이 결과표에 개별 행(Row)으로 나타나야 합니다.
                3. 예를 들어 '복리후생비', '접대비', '통신비'를 '판관비' 하나로 퉁치지 말고, 각각의 행으로 모두 나열하십시오.
                4. 각 계정의 상위 분류(예: 유동자산, 비유동부채, 매출원가, 판관비 등)를 'Class' 컬럼에 명시하십시오.

                [출력 포맷]
                결과는 오직 **JSON 배열** 형태여야 합니다.
                JSON 구조:
                [
                  {{
                    "Class": "판매비와관리비",
                    "Account_Name": "복리후생비",
                    "2022": 1500000,
                    "2023": 1600000,
                    "2024": 0
                  }},
                  ...
                ]

                [데이터 처리 규칙]
                1. 연도(Year)는 데이터 내의 날짜나 컬럼명을 보고 2022, 2023, 2024 등으로 자동 할당하십시오.
                2. 금액은 정확히 합산하십시오.
                3. 값이 없는 연도는 0으로 표기하십시오.

                [분석할 데이터]:
                {csv_data}
                """
                
                # --- [요청사항 반영] 모델 고정 ---
                response = client.models.generate_content(
                    model="gemini-3-flash-preview", 
                    contents=prompt
                )
                
                # 3. 결과 처리
                try:
                    cleaned_text = response.text.replace("```json", "").replace("```", "").strip()
                    # 가끔 JSON 앞뒤에 설명이 붙는 경우를 대비해 '['와 ']' 사이만 추출 시도
                    if "[" in cleaned_text and "]" in cleaned_text:
                        start_idx = cleaned_text.find("[")
                        end_idx = cleaned_text.rfind("]") + 1
                        cleaned_text = cleaned_text[start_idx:end_idx]

                    ai_result_json = json.loads(cleaned_text)
                    ai_df = pd.DataFrame(ai_result_json)
                    
                    # 컬럼 순서 보기 좋게 정렬 (Class, Account_Name 먼저)
                    cols = ai_df.columns.tolist()
                    front_cols = ['Class', 'Account_Name']
                    year_cols = sorted([c for c in cols if c not in front_cols])
                    final_cols = front_cols + year_cols
                    # 실제 존재하는 컬럼만 선택
                    final_cols = [c for c in final_cols if c in ai_df.columns]
                    ai_df = ai_df[final_cols]

                    st.subheader("🏆 상세 재무제표 (전체 계정)")
                    st.dataframe(ai_df, use_container_width=True)
                    
                    # 4. 엑셀 다운로드
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        ai_df.to_excel(writer, sheet_name='Financial_Statements', index=False)
                        merged_df.to_excel(writer, sheet_name='Raw_Data', index=False)
                    
                    st.download_button(
                        label="📥 상세 재무제표 엑셀 다운로드",
                        data=buffer.getvalue(),
                        file_name="financial_statements_detail.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except json.JSONDecodeError:
                    st.error("데이터가 너무 많거나 AI 응답 형식이 올바르지 않습니다. 원본 응답을 확인하세요.")
                    st.text_area("AI 응답", response.text, height=300)
                    
        except Exception as e:
            st.error(f"오류가 발생했습니다: {e}")
            # 모델명 에러일 경우 힌트 제공
            if "404" in str(e) or "not found" in str(e).lower():
                st.warning("⚠️ 'gemini-3-flash-preview' 모델을 찾을 수 없습니다. 사용 가능한 모델명인지 확인하거나, 코드를 'gemini-1.5-flash' 등으로 변경해보세요.")
        finally:
            progress_text.empty()