import streamlit as st
import pandas as pd
import io
import re
import logic 
from google import genai # 채팅용 클라이언트 직접 호출을 위해 추가
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# 페이지 설정
st.set_page_config(page_title="Financial Report AI", layout="wide")

# --- CSS: 스타일링 ---
st.markdown("""
    <style>
        .file-list-box {
            border: 1px solid #e6e6e6; padding: 10px; border-radius: 5px;
            max-height: 200px; overflow-y: auto; background-color: #f9f9f9; margin-bottom: 20px;
        }
        .file-item {
            font-size: 0.9em; margin-bottom: 4px; padding: 4px; background: white; border-radius: 3px;
        }
        div[data-testid="stVerticalBlock"] > div[style*="flex-direction: column;"] > div[data-testid="stVerticalBlock"] {
            gap: 0rem;
        }
    </style>
""", unsafe_allow_html=True)

# --- 초기화 콜백 함수 ---
def clear_all():
    # 파일 업로더 초기화
    if 'uploader_key' in st.session_state:
        st.session_state['uploader_key'] = []
    # 데이터 초기화
    if 'raw_data' in st.session_state:
        del st.session_state['raw_data']
    # [추가] 채팅 기록 초기화
    if 'messages' in st.session_state:
        del st.session_state['messages']

# --- 컬럼 정렬 함수 ---
def sort_columns_chronologically(columns):
    fixed_cols = ['Account_Name']
    date_cols = [c for c in columns if c not in ['Statement', 'Level', 'Account_Name']]
    
    def date_sort_key(col_name):
        s_name = str(col_name)
        year_match = re.search(r'(\d{4})', s_name)
        year = int(year_match.group(1)) if year_match else 9999
        
        sub_val = 0
        if '1Q' in s_name: sub_val = 1
        elif '2Q' in s_name: sub_val = 4
        elif '3Q' in s_name: sub_val = 7
        elif '4Q' in s_name: sub_val = 10
        
        is_cum = 1 if '누적' in s_name or 'Cum' in s_name or 'Year' in s_name else 0
        return (year, sub_val, is_cum, s_name)
    
    sorted_date_cols = sorted(date_cols, key=date_sort_key)
    return fixed_cols + sorted_date_cols

# --- 스타일 함수 ---
def style_dataframe(row):
    level = row.get('Level', 3)
    if level == 1: return ['background-color: #1f77b4; color: white; font-weight: bold;'] * len(row)
    elif level == 2: return ['background-color: #aec7e8; color: black; font-weight: bold;'] * len(row)
    return ['color: black;'] * len(row)

# --- 엑셀 저장 함수 ---
def save_styled_excel(df, sheet_name_map, unit_text):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        if 'Statement' in df.columns: statements = df['Statement'].unique()
        else: statements = ['Result']
            
        for stmt in statements:
            if 'Statement' in df.columns: sub_df = df[df['Statement'] == stmt].copy()
            else: sub_df = df.copy()
            
            all_cols = sub_df.columns.tolist()
            sorted_cols = sort_columns_chronologically(all_cols)
            final_cols = [c for c in sorted_cols if c in sub_df.columns]
            
            sheet_title = sheet_name_map.get(stmt, stmt)[:30]
            sub_df[final_cols].to_excel(writer, sheet_name=sheet_title, index=False, startrow=1)
            
            ws = writer.sheets[sheet_title]
            ws['A1'] = f"(단위: {unit_text})"
            ws['A1'].font = Font(bold=True, italic=True)
            
            fill_lv1 = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
            font_lv1 = Font(color="FFFFFF", bold=True)
            fill_lv2 = PatternFill(start_color="AEC7E8", end_color="AEC7E8", fill_type="solid")
            font_lv2 = Font(color="000000", bold=True)
            
            numeric_col_indices = [i+1 for i, c in enumerate(final_cols) if c != 'Account_Name']
            
            sub_df = sub_df.reset_index(drop=True)
            for idx, row in sub_df.iterrows():
                excel_row = idx + 3
                level = row.get('Level', 3)
                for col_idx in range(1, len(final_cols) + 1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if level == 1:
                        cell.fill = fill_lv1
                        cell.font = font_lv1
                    elif level == 2:
                        cell.fill = fill_lv2
                        cell.font = font_lv2
                    if col_idx - 1 in numeric_col_indices:
                        cell.number_format = '#,##0'
            ws.column_dimensions['A'].width = 30
    return buffer

# ==========================================
# [UI 1] 타이틀 및 API Key
# ==========================================
st.title("📑 통합 재무제표 보고서")

if 'api_key' not in st.session_state:
    st.session_state.api_key = ''

api_key_input = st.text_input(
    "Gemini API Key", 
    type="password", 
    placeholder="sk-...", 
    value=st.session_state.api_key
)
if api_key_input:
    st.session_state.api_key = api_key_input

st.divider()

# ==========================================
# [UI 2] 파일 업로더 & 초기화
# ==========================================
col_upload, col_clear = st.columns([0.85, 0.15])

with col_upload:
    uploaded_files = st.file_uploader(
        "분석할 파일들을 업로드하세요", 
        accept_multiple_files=True, 
        type=['xlsx', 'xls', 'csv', 'pdf', 'docx', 'txt'],
        key='uploader_key'
    )

with col_clear:
    st.write("") 
    st.write("")
    if st.button("🗑️ 초기화", type="secondary", use_container_width=True, on_click=clear_all):
        pass

if uploaded_files:
    file_list_html = '<div class="file-list-box">'
    for f in uploaded_files:
        size_kb = f.size / 1024
        file_list_html += f'<div class="file-item">📄 {f.name} ({size_kb:.1f} KB)</div>'
    file_list_html += '</div>'
    st.markdown(file_list_html, unsafe_allow_html=True)

    if st.session_state.api_key:
        if st.button("🚀 보고서 생성 시작", type="primary", use_container_width=True):
            status = st.status("AI가 분석 중입니다...", expanded=True)
            try:
                raw_df = logic.process_smart_merge(st.session_state.api_key, uploaded_files)
                for col in raw_df.columns:
                    if col not in ['Statement', 'Level', 'Account_Name']:
                        raw_df[col] = pd.to_numeric(raw_df[col], errors='coerce').fillna(0)
                
                numeric_cols = [c for c in raw_df.columns if c not in ['Statement', 'Level', 'Account_Name']]
                zero_cols = [c for c in numeric_cols if raw_df[c].abs().sum() == 0]
                if zero_cols:
                    raw_df = raw_df.drop(columns=zero_cols)
                
                st.session_state['raw_data'] = raw_df
                # 분석 새로 하면 채팅 기록도 리셋
                if 'messages' in st.session_state:
                    del st.session_state['messages']
                    
                status.update(label="✅ 분석 완료!", state="complete", expanded=False)
            except Exception as e:
                status.update(label="❌ 오류 발생", state="error")
                st.error(f"에러 내용: {e}")
    else:
        st.warning("👆 상단에 API Key를 먼저 입력해주세요.")


# ==========================================
# [UI 3] 분석 결과 & 채팅 인터페이스
# ==========================================
if 'raw_data' in st.session_state:
    st.divider()
    
    # 1. 상단: 테이블 뷰
    c_title, c_unit = st.columns([0.7, 0.3])
    with c_unit:
        unit_option = st.selectbox("단위 선택", ("원", "천원", "백만원", "억원"), index=0)
        unit_divisors = {"원": 1, "천원": 1000, "백만원": 1000000, "억원": 100000000}
        divisor = unit_divisors[unit_option]

    with c_title:
        st.subheader(f"📊 분석 결과 (단위: {unit_option})")

    display_df = st.session_state['raw_data'].copy()
    numeric_cols = [c for c in display_df.columns if c not in ['Statement', 'Level', 'Account_Name']]
    display_df = display_df[display_df[numeric_cols].abs().sum(axis=1) != 0]
    
    for col in numeric_cols:
        if divisor > 1:
            display_df[col] = display_df[col] / divisor

    available_types = display_df['Statement'].unique() if 'Statement' in display_df.columns else []
    type_map = {
        'BS': '재무상태표', 'IS': '손익계산서', 'COGM': '제조원가명세서', 
        'CF': '현금흐름표', 'SCE': '자본변동표', 'RE': '이익잉여금', 'Other': '기타'
    }
    
    if len(available_types) > 0:
        tabs = st.tabs([type_map.get(t, t) for t in available_types])
        for i, stmt_type in enumerate(available_types):
            with tabs[i]:
                sub_df = display_df[display_df['Statement'] == stmt_type].copy()
                all_cols = sub_df.columns.tolist()
                sorted_cols = sort_columns_chronologically(all_cols)
                final_cols = [c for c in sorted_cols if c in sub_df.columns]
                format_dict = {col: "{:,.0f}" for col in numeric_cols if col in final_cols}
                st.dataframe(sub_df[final_cols].style.apply(style_dataframe, axis=1).format(format_dict), use_container_width=True, height=600)
    
    excel_buffer = save_styled_excel(display_df, type_map, unit_option)
    st.download_button(
        f"📥 엑셀 다운로드 (현재 단위: {unit_option})",
        data=excel_buffer.getvalue(),
        file_name=f"Financial_Report_{unit_option}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ==========================================
    # [NEW] 2. 하단: AI 추가 요청 채팅창
    # ==========================================
    st.divider()
    st.subheader("💬 AI 재무 비서")
    st.info(f"위의 **분석된 데이터({unit_option} 단위)**를 바탕으로 추가 질문을 하거나 분석을 요청할 수 있습니다.")

    # 세션 스테이트에 대화 기록 초기화
    if "messages" not in st.session_state:
        st.session_state["messages"] = []

    # 이전 대화 내용 출력
    for msg in st.session_state["messages"]:
        st.chat_message(msg["role"]).write(msg["content"])

    # 사용자 입력 처리
    if prompt := st.chat_input("예: 2024년 영업이익률은 얼마인가요? / 매출액 추이를 요약해줘"):
        # 1. 사용자 메시지 표시
        st.session_state["messages"].append({"role": "user", "content": prompt})
        st.chat_message("user").write(prompt)

        # 2. AI 응답 생성
        # 현재 보고 있는 데이터프레임(display_df)을 CSV 텍스트로 변환하여 컨텍스트로 전달
        context_csv = display_df.to_csv(index=False)
        
        # 시스템 프롬프트 구성
        system_prompt = f"""
        당신은 유능한 재무 분석가입니다. 
        사용자는 아래의 재무제표 데이터(CSV 포맷, 단위: {unit_option})를 보고 있습니다.
        사용자의 질문에 대해 데이터를 기반으로 명확하고 통찰력 있게 답변하세요.
        
        [데이터]
        {context_csv}
        
        [답변 가이드]
        - 구체적인 수치를 인용하세요.
        - 추세나 특이사항이 있다면 언급하세요.
        - 표나 마크다운 형식을 적절히 사용하세요.
        """

        try:
            client = genai.Client(api_key=st.session_state.api_key)
            # 채팅 히스토리 없이 1회성 질문/답변 구조 (데이터가 크므로 매번 컨텍스트 주입이 유리)
            response = client.models.generate_content(
                model="gemini-3-flash-preview",
                contents=f"{system_prompt}\n\n[사용자 질문]: {prompt}"
            )
            ai_reply = response.text
            
            # 3. AI 메시지 표시
            st.session_state["messages"].append({"role": "assistant", "content": ai_reply})
            st.chat_message("assistant").write(ai_reply)
            
        except Exception as e:
            st.error(f"답변 생성 중 오류가 발생했습니다: {e}")